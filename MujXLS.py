# -*- coding: cp1250 -*-
# verze 2018-03
# created by Milan Hutera
# working with excel files


from numpy import append
import openpyxl
import types
from operator import itemgetter
from copy import copy
from copy import deepcopy


class XLS:
    def sortColumnDimenstion(self, SpreadSheet):
        lastWidth = 3.6
        for Sheet in SpreadSheet:
            if 'ColumnDimension' in SpreadSheet[Sheet]:
                SpreadSheet[Sheet]['ColumnDimension'] = deepcopy(SpreadSheet[Sheet]['ColumnDimension'])
            for Column in SpreadSheet[Sheet]['Head']:
                if Column in SpreadSheet[Sheet]['ColumnDimension']:                    
                    Index = chr(64 + SpreadSheet[Sheet]['Head'][Column])
                    if Index != SpreadSheet[Sheet]['ColumnDimension'][Column].index:
                        print(Sheet,Column, Index, SpreadSheet[Sheet]['ColumnDimension'][Column].index)
                        print (SpreadSheet[Sheet]['Head'][Column])
                    if SpreadSheet[Sheet]['ColumnDimension'][Column].width is None:
                        SpreadSheet[Sheet]['ColumnDimension'][Column].width = lastWidth
                    else:
                        lastWidth = SpreadSheet[Sheet]['ColumnDimension'][Column].width
                        
                    SpreadSheet[Sheet]['ColumnDimension'][Column].index = Index
                    SpreadSheet[Sheet]['ColumnDimension'][Column].min = SpreadSheet[Sheet]['Head'][Column]
                    SpreadSheet[Sheet]['ColumnDimension'][Column].max = SpreadSheet[Sheet]['Head'][Column]
        return SpreadSheet
                



                                             
    def sheetList(self, Soubor):
        try:
            excel_document = openpyxl.load_workbook(Soubor)
            return excel_document.sheetnames
        except:
            return []
        
    
    def __HeadXLS(self, Source,Row = 1):
        Index = 1
        Head = {}
        for i in range(1,Source.max_column+1):
            if Source.cell(row=Row,column=i).value!= None:
                if (Source.cell(row=Row,column=i).value in Head)==False:
                    Aux = str(Source.cell(row=Row,column=i).value)                    
                    Head[Aux]=Index
                else:
                    Head[str(Index)] = Index
            else:
                Head[str(Index)] = Index
            Index = Index + 1
        
        return Head
    def __CopyStyle(self, Source, Head):
        StyleData = {}
        StyleHead = {}
        ColumnDimension = {}
        for i in Head:
            Column = Source.cell(row=1,column=Head[i]).column
            StyleHead[i] = Source.cell(row=1,column=Head[i])
            ColumnDimension[i] = Source.column_dimensions[Column]
            StyleData[i] = Source.cell(row=2,column=Head[i])
            
        return StyleHead, StyleData, ColumnDimension
            
    
    def __DataXLS(self, Source, Head, Row = 2):
        Data = []
        for i in range(Row,Source.max_row + 1 ):
            Aux = {}
            Add = False
            for ii in Head:
                
                Aux[ii] = Source.cell(row=i,column=Head[ii]).value
                if Aux[ii]!=None:
                    Add = True
                else:
                    Aux[ii] = ''
            if Add:
                Data.append(Aux)
                
        return Data

    def __DataXLSstr(self, Source, Head, Data = [], Row = 2):
        for i in range(Row,Source.max_row + 1 ):
            Aux = {}
            Add = False
            for ii in Head:
                
                Aux[ii] = Source.cell(row=i,column=Head[ii]).value
                if Aux[ii]!=None:
                    Add = True
                else:
                    Aux[ii] = ''
                Aux[ii] = str(Aux[ii])
            if Add:
                Data.append(Aux)
                
        return Data
    def RefSheets(self, Soubor, Sheet, ColumnList):
        ColumnListUse = []
        excel = self.ReadXLS(Soubor)
        head = excel[Sheet]['Head']

        if ColumnList[0] in head:
            for Column in ColumnList:
                ColumnListUse.append(head[Column])
        else:
            for Column in ColumnList:                    
                if not(type(Column) == int):
                    ColumnListUse.append( ord(Column.upper())-64)
                else:
                    ColumnListUse.append(Column)
                
        location = {}
        excel_document = openpyxl.load_workbook(Soubor)
        sheet = excel_document[Sheet]
        for i in range(2,sheet.max_row + 1 ):
            key = ''
            for Column in ColumnListUse:
                if type(sheet.cell(row=i,column=Column).value)!=type(None):  
                    key += sheet.cell(row=i,column=Column).value
            if key != '':
                if not(key in location):
                    location[key] = Sheet + "!" + chr(64 + ColumnListUse[0])+str(i)
                else:
                    location[key] = location[key].split(":")[0] + ":" + chr(64 + ColumnListUse[0])+str(i)
        excel_document.close()
        return location

     
    
    def SetHyperlink(self, Soubor, Sheet, ColumnList, Links):
        ColumnListUse = []        
        for Column in ColumnList:                    
            if not(type(Column) == int):
                ColumnListUse.append( ord(Column.upper())-64)
            else:
                ColumnListUse.append(Column)

        tempHyp = openpyxl.worksheet.hyperlink.Hyperlink
        tempHyp.ref = ""
            
        excel_document = openpyxl.load_workbook(Soubor)
        sheet = excel_document[Sheet]
        for i in range(2,sheet.max_row + 1 ):            
            key = ''
            for Column in ColumnListUse:
                
                if type(sheet.cell(row=i,column=Column).value)!=type(None):
                    
                    key += sheet.cell(row=i,column=Column).value
                    
            if key in Links:
                hyperlink = tempHyp()
                hyperlink.location = Links[key]
                sheet.cell(row=i,column=ColumnListUse[0]).hyperlink = hyperlink
                
        
        excel_document.save(Soubor)
        excel_document.close()

    def SetRows(self, Soubor, Sheet, ColumnList, Color = "DDDDDD", Color2 = "FFFFFF"):
        ColumnListUse = []        
        for Column in ColumnList:                    
            if not(type(Column) == int):
                ColumnListUse.append( ord(Column.upper())-64)
            else:
                ColumnListUse.append(Column)
                
        excel_document = openpyxl.load_workbook(Soubor)
        sheet = excel_document[Sheet]
        old = ''
        for Column in ColumnListUse:
            if type(sheet.cell(row=2,column=Column).value)!=type(None):  
                old += sheet.cell(row=2,column=Column).value
                
        color = False
        fill = openpyxl.styles.PatternFill("solid", fgColor=Color)
        fill_2 = openpyxl.styles.PatternFill("solid", fgColor=Color2)
        for i in range(2,sheet.max_row + 1 ):
            current = ''
            for Column in ColumnListUse:
                if type(sheet.cell(row=i,column=Column).value)!=type(None):  
                    current += sheet.cell(row=i,column=Column).value

                
            if old != current:
                color = not(color)
                old = current
            if color:
                for c in range(1,sheet.max_column+1):
                    cell = sheet.cell(row=i,column=c)
                    cell.fill = fill
            else:
                for c in range(1,sheet.max_column+1):
                    cell = sheet.cell(row=i,column=c)
                    cell.fill = fill_2
        excel_document.save(Soubor)
        excel_document.close()

    def NewSheet(self, SheetName):
        result = {}
        result[SheetName] = {}
        result[SheetName]['Head'] = {}
        result[SheetName]['Data'] = []
        return result
                               
    
    def ReadXLS(self, Soubor, Sheets = [], Head = {}, HeadRow = 1, DataRow = 2):
        workbook = {}
        excel_document = openpyxl.load_workbook(Soubor)
        
        
        if Sheets == []:
            Sheets = excel_document.sheetnames
        for i in  Sheets:
            workbook[i] = {}
            workbook[i]['Head'] = {}
            workbook[i]['Data'] = []
            
##            sheet = excel_document.get_sheet_by_name(i)
            sheet = excel_document[i]
            if Head == {}:
                workbook[i]['Head'] = self.__HeadXLS(sheet, Row = HeadRow)
            else:
                workbook[i]['Head'] = Head
            workbook[i]['Data'] = self.__DataXLS(sheet, workbook[i]['Head'], Row = DataRow )
            workbook[i]['HeadStyle'], workbook[i]['DataStyle'], workbook[i]['ColumnDimension'] = self.__CopyStyle(sheet, workbook[i]['Head'])
        excel_document.close()
        return workbook
    
    def ReadXLSstr(self, Soubor, Sheets = [], Head = {}, HeadRow = 1, DataRow = 2):
        workbook = {}
        excel_document = openpyxl.load_workbook(Soubor)
        if Sheets == []:
            Sheets = excel_document.sheetnames
        for i in  Sheets:
            workbook[i] = {}
            workbook[i]['Head'] = {}
            workbook[i]['Data'] = []
            
##            sheet = excel_document.get_sheet_by_name(i)
            sheet = excel_document[i]
            if Head == {}:
                workbook[i]['Head'] = self.__HeadXLS(sheet, Row = HeadRow)
            else:
                workbook[i]['Head'] = Head
            workbook[i]['Data'] = self.__DataXLSstr(sheet, workbook[i]['Head'], workbook[i]['Data'], Row = DataRow )
        excel_document.close()
        return workbook
    def CSVHead(self, Head):
        aux = list(Head.values())
        aux.sort()
        result = list(range(aux[-1]))
        for i in Head:
            result[Head[i] - 1] = i
        return result

    def createHead(self, Head):
        """from array creates a Json """
        count = 1
        result = {}
        for i in Head:
            result[i] = count
            count += 1
        return result

    def SaveCSV(self, Workbook, Sheet):
        
        lastHead = 0
        Head = list(range(len(Workbook[Sheet]['Head'])))
        for columnName in Workbook[Sheet]['Head']:
            column = Workbook[Sheet]['Head'][columnName] - 1
            Head[column] = columnName
                    

        Data = ''
        for row in Workbook[Sheet]['Data']:
            newRow = [""] * len(Workbook[Sheet]['Head'])
            for columnName in row:
                if columnName in Workbook[Sheet]['Head']:
                    column = Workbook[Sheet]['Head'][columnName] - 1
                    newRow[column] = row[columnName]
                else:
                    if lastHead == 0:
                        for headItem in Workbook[row]['Head']:
                            lastHead = max(lastHead,Workbook[i]['Head'][headItem])
                    lastHead += 1
                    Workbook[i]['Head'][columnName] = lastHead
                    Head.append(columnName)
                    newRow.append(row[columnName])
            Data += str(newRow)[1:-1]
            Data += "\r\n"
        
        Data = str(Head)[1:-1] + "\r\n" + Data
        return Data
        

    def SaveXLS(self, FileName, Workbook, Sheets = []):
        
        wb = openpyxl.Workbook()
        DefaultSheet = wb.sheetnames[0]
        lastHead = 0
        if Sheets == []:
            Sheets = list(Workbook.keys())
        for Sheet in Sheets:
            if Sheets.index(Sheet) == 0 and wb.sheetnames.count(DefaultSheet)>0:
                #MySheet = wb.get_sheet_by_name(DefaultSheet)
                MySheet = wb[DefaultSheet]
                MySheet.title = Sheet
            else:
                wb.create_sheet(Sheet)
                #MySheet = wb.get_sheet_by_name(Sheet)
                MySheet = wb[Sheet]
            for HeadItem in Workbook[Sheet]['Head']:
                MySheet.cell(row=1,column=Workbook[Sheet]['Head'][HeadItem]).value = HeadItem
            if 'ColumnDimension' in Workbook[Sheet]:
                for ColumnDimenstionItem in Workbook[Sheet]['ColumnDimension']:
                    Column = MySheet.cell(row=1,column=Workbook[Sheet]['Head'][ColumnDimenstionItem]).column
                    Workbook[Sheet]['ColumnDimension'][ColumnDimenstionItem].index = Column
                    Workbook[Sheet]['ColumnDimension'][ColumnDimenstionItem].max = Workbook[Sheet]['Head'][ColumnDimenstionItem]
                    Workbook[Sheet]['ColumnDimension'][ColumnDimenstionItem].min = Workbook[Sheet]['Head'][ColumnDimenstionItem]
                    MySheet.column_dimensions[Column] = Workbook[Sheet]['ColumnDimension'][ColumnDimenstionItem]

            if 'HeadStyle' in Workbook[Sheet]:
                for HeadStyleItem in Workbook[Sheet]['HeadStyle']:
                    MySheet.cell(row=1,column=Workbook[Sheet]['Head'][HeadStyleItem]).font           = copy(Workbook[Sheet]['HeadStyle'][HeadStyleItem].font)
                    MySheet.cell(row=1,column=Workbook[Sheet]['Head'][HeadStyleItem]).border         = copy(Workbook[Sheet]['HeadStyle'][HeadStyleItem].border)
                    MySheet.cell(row=1,column=Workbook[Sheet]['Head'][HeadStyleItem]).fill           = copy(Workbook[Sheet]['HeadStyle'][HeadStyleItem].fill)
                    MySheet.cell(row=1,column=Workbook[Sheet]['Head'][HeadStyleItem]).number_format  = copy(Workbook[Sheet]['HeadStyle'][HeadStyleItem].number_format)
                    MySheet.cell(row=1,column=Workbook[Sheet]['Head'][HeadStyleItem]).protection     = copy(Workbook[Sheet]['HeadStyle'][HeadStyleItem].protection)
                    MySheet.cell(row=1,column=Workbook[Sheet] ['Head'][HeadStyleItem]).alignment      = copy(Workbook[Sheet]['HeadStyle'][HeadStyleItem].alignment)                
                    
                    
                    
            row = 2
            for DataRow in Workbook[Sheet]['Data']:
                for DataItem in DataRow:
                    if DataItem in Workbook[Sheet]['Head']:
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataItem]).value = DataRow[DataItem]
                    else:
                        if lastHead == 0:
                            for headItem in Workbook[Sheet]['Head']:
                                lastHead = max(lastHead,Workbook[Sheet]['Head'][headItem])
                        lastHead += 1
                        Workbook[Sheet]['Head'][DataItem] = lastHead
                        MySheet.cell(row=1,column=Workbook[Sheet]['Head'][DataItem]).value = DataItem
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataItem]).value = DataRow[DataItem]
                if 'DataStyle' in Workbook[Sheet]:
                    for DataStyleItem in Workbook[Sheet]['DataStyle']:                        
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataStyleItem]).font           = copy(Workbook[Sheet]['DataStyle'][DataStyleItem].font)
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataStyleItem]).border         = copy(Workbook[Sheet]['DataStyle'][DataStyleItem].border)
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataStyleItem]).fill           = copy(Workbook[Sheet]['DataStyle'][DataStyleItem].fill)
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataStyleItem]).number_format  = copy(Workbook[Sheet]['DataStyle'][DataStyleItem].number_format)
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataStyleItem]).protection     = copy(Workbook[Sheet]['DataStyle'][DataStyleItem].protection)
                        MySheet.cell(row=row,column=Workbook[Sheet]['Head'][DataStyleItem]).alignment     = copy(Workbook[Sheet]['DataStyle'][DataStyleItem].alignment)
                    
                row += 1
        print ("save", FileName)
        wb.save(FileName)
        wb.close()

if __name__ == '__main__':
    result = []
    import MujXLS
    zz = MujXLS.XLS()
    soubor = 'c:\\Users\\milanhutera\\OneDrive - Doosan\\Python\\FaultTable\\ddd1.xlsx'
    wb = zz.ReadXLS('c:\\Users\\milanhutera\\OneDrive - Doosan\\Python\\FaultTable\\ddd1.xlsx')
    zz.SaveXLS('c:\\Users\\milanhutera\\OneDrive - Doosan\\Python\\FaultTable\\ddd2.xlsx',wb)
